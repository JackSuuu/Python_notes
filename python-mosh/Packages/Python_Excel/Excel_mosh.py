import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("/Users/jack/Desktop/Python/Py_P_Package/Python_Excel/2022_CIE_seating_plan.xlsx")
print(wb.sheetnames)

sheet = wb["May 30-Jun 3"]
# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

cell = sheet["a654"]
# cell = sheet.cell(row=654, column=1)
# print(cell.value)
# # cell.value = 1
# print(cell.row)
# print(cell.column)
print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, column=4)
        if cell.value == "王秭":
            print("The location of my name is:", row)

cells = sheet["a:c"]
column = sheet["a"]
cells = sheet["a1:c3"]
print(cells)

sheet.append([1, 2, 3])
wb.save("New_excel_file")
sheet.insert_cols
sheet.delete.cols


